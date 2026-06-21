import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Define file paths
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
DATA_FILE = os.path.join(PROJECT_ROOT, 'data', 'banks_q3fy26.json')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'reports')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'bank_soundness_report_q3fy26.xlsx')

# RBI PCA minimum values
CRAR_MIN = 11.50
CET1_MIN = 8.00
NETNPA_TRIGGER = 6.00

# D-SIB Tickers
D_SIBS = ['SBIN', 'HDFCBANK', 'ICICIBANK']

def is_d_sib(ticker):
    return ticker.upper() in D_SIBS

def assess_crar(val):
    if val is None:
        return 'PENDING', 'Data Pending', None
    headroom_bps = int(round((val - CRAR_MIN) * 100))
    if val >= CRAR_MIN:
        return 'CLEAR', f'Clear (+{headroom_bps} bps)', headroom_bps
    elif val >= 9.00:
        return 'THRESHOLD_1', f'Risk 1 (-{abs(headroom_bps)} bps)', headroom_bps
    elif val >= 7.50:
        return 'THRESHOLD_2', f'Risk 2 (-{abs(headroom_bps)} bps)', headroom_bps
    else:
        return 'THRESHOLD_3', f'Risk 3 (-{abs(headroom_bps)} bps)', headroom_bps

def assess_cet1(val):
    if val is None:
        return 'PENDING', 'Data Pending', None
    headroom_bps = int(round((val - CET1_MIN) * 100))
    if val >= CET1_MIN:
        return 'CLEAR', f'Clear (+{headroom_bps} bps)', headroom_bps
    elif val >= 5.50:
        return 'THRESHOLD_1', f'Risk 1 (-{abs(headroom_bps)} bps)', headroom_bps
    elif val >= 4.00:
        return 'THRESHOLD_2', f'Risk 2 (-{abs(headroom_bps)} bps)', headroom_bps
    else:
        return 'THRESHOLD_3', f'Risk 3 (-{abs(headroom_bps)} bps)', headroom_bps

def assess_net_npa(val):
    if val is None:
        return 'PENDING', 'Data Pending', None
    headroom_bps = int(round((6.00 - val) * 100))
    if val < 6.00:
        return 'CLEAR', f'Clear (+{headroom_bps} bps)', headroom_bps
    elif val < 9.00:
        return 'THRESHOLD_1', f'Risk 1 (-{abs(headroom_bps)} bps)', headroom_bps
    elif val < 12.00:
        return 'THRESHOLD_2', f'Risk 2 (-{abs(headroom_bps)} bps)', headroom_bps
    else:
        return 'THRESHOLD_3', f'Risk 3 (-{abs(headroom_bps)} bps)', headroom_bps

def assess_leverage(val, ticker):
    min_req = 4.00 if is_d_sib(ticker) else 3.50
    if val is None:
        return 'PENDING', 'Data Pending', None
    headroom_bps = int(round((val - min_req) * 100))
    if val >= min_req:
        return 'CLEAR', f'Clear (+{headroom_bps} bps)', headroom_bps
    elif abs(headroom_bps) <= 50:
        return 'THRESHOLD_1', f'Risk 1 (-{abs(headroom_bps)} bps)', headroom_bps
    elif abs(headroom_bps) <= 100:
        return 'THRESHOLD_2', f'Risk 2 (-{abs(headroom_bps)} bps)', headroom_bps
    else:
        return 'THRESHOLD_3', f'Risk 3 (-{abs(headroom_bps)} bps)', headroom_bps

def grade_to_score(grade):
    scores = {'CLEAR': 0, 'PENDING': 0, 'THRESHOLD_1': 1, 'THRESHOLD_2': 2, 'THRESHOLD_3': 3}
    return scores.get(grade, 0)

def get_overall_grade(crar_g, cet1_g, npa_g, lev_g):
    # Capital status is the worse of CRAR and CET1
    cap_grade = crar_g if grade_to_score(crar_g) > grade_to_score(cet1_g) else cet1_g
    
    cap_s = grade_to_score(cap_grade)
    aq_s = grade_to_score(npa_g)
    lev_s = grade_to_score(lev_g)
    
    scores = [cap_s, aq_s, lev_s]
    max_s = max(scores)
    
    # Rule check: Threshold 3 on any, or Threshold 2 on 2 or more
    count_t2 = len([s for s in scores if s >= 2])
    
    if max_s == 3 or count_t2 >= 2:
        return 'THRESHOLD_3'
    elif max_s == 2:
        return 'THRESHOLD_2'
    elif max_s == 1:
        return 'THRESHOLD_1'
    return 'CLEAR'

def main():
    print(f"Reading bank metrics from {DATA_FILE}...")
    if not os.path.exists(DATA_FILE):
        print(f"Error: Data file not found at {DATA_FILE}")
        return

    with open(DATA_FILE, 'r') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "PCA Soundness Monitor"
    ws.views.sheetView[0].showGridLines = True

    # Typography & Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F3A5F")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="565F6B")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    data_bold = Font(name=font_family, size=10, bold=True)
    disclaimer_font = Font(name=font_family, size=8, italic=True, color="565F6B")

    # Header Fills
    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    
    # Risk Fills (Soft Colors)
    fill_clear = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    fill_watch = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
    fill_breach = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red/Orange
    fill_pending = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Soft Gray

    # Cell Borders
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    double_border_side = Side(border_style="double", color="1F3A5F")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    summary_border = Border(top=thin_border_side, bottom=double_border_side)

    # Title Block
    ws['A1'] = "Bank Soundness Early-Warning Monitor"
    ws['A1'].font = title_font
    ws['A2'] = f"Quarter: {data['quarter']} | Period Ended: {data['periodEnded']} | Currency: {data['currency']} | Framework: RBI Circular RBI/2021-22/118"
    ws['A2'].font = subtitle_font
    
    # Spacer
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 10

    # Table Headers
    headers = [
        "Bank Entity", "Ticker", "Sector", 
        "CRAR (%)", "CRAR Status", 
        "CET1 (%)", "CET1 Status", 
        "Net NPA (%)", "Net NPA Status", 
        "Leverage (%)", "Leverage Status", 
        "Overall PCA Grade"
    ]
    
    header_row = 5
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'))
    ws.row_dimensions[header_row].height = 28

    # Write Data
    start_row = 6
    for idx, bank_raw in enumerate(data['banks']):
        row = start_row + idx
        ticker = bank_raw.get('ticker', '')
        
        # Assess Ratios
        crar_val = bank_raw.get('crar')
        crar_grade, crar_desc, _ = assess_crar(crar_val)
        
        cet1_val = bank_raw.get('cet1')
        cet1_grade, cet1_desc, _ = assess_cet1(cet1_val)
        
        npa_val = bank_raw.get('netNpaRatio')
        npa_grade, npa_desc, _ = assess_net_npa(npa_val)
        
        # Leverage is missing in Q3 FY26, check if future proof exists
        lev_val = bank_raw.get('leverageRatio')
        lev_grade, lev_desc, _ = assess_leverage(lev_val, ticker)
        
        overall_g = get_overall_grade(crar_grade, cet1_grade, npa_grade, lev_grade)

        # Write Core Columns
        ws.cell(row=row, column=1, value=bank_raw.get('name')).font = data_bold
        ws.cell(row=row, column=2, value=ticker).font = data_font
        ws.cell(row=row, column=3, value=bank_raw.get('type')).font = data_font
        
        # Format alignment of metadata
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

        # Ratios writing function
        def write_metric(col_idx, val, grade, desc):
            c_val = ws.cell(row=row, column=col_idx)
            c_desc = ws.cell(row=row, column=col_idx+1)
            
            c_val.font = data_font
            c_desc.font = data_font
            c_val.border = cell_border
            c_desc.border = cell_border
            c_val.alignment = Alignment(horizontal="right", vertical="center")
            c_desc.alignment = Alignment(horizontal="center", vertical="center")
            
            if val is not None:
                c_val.value = val / 100.0 # openpyxl format percentage requires decimal divided by 100
                c_val.number_format = '0.00%'
            else:
                c_val.value = "Data Pending"
                c_val.alignment = Alignment(horizontal="center", vertical="center")
                c_val.font = Font(name=font_family, size=10, color="8A8F98")

            c_desc.value = desc
            
            # Apply fill color based on risk grade
            fill_map = {
                'CLEAR': fill_clear,
                'THRESHOLD_1': fill_watch,
                'THRESHOLD_2': fill_breach,
                'THRESHOLD_3': fill_breach,
                'PENDING': fill_pending
            }
            c_desc.fill = fill_map.get(grade, fill_pending)
            if grade == 'PENDING':
                c_val.fill = fill_pending

        write_metric(4, crar_val, crar_grade, crar_desc)
        write_metric(6, cet1_val, cet1_grade, cet1_desc)
        write_metric(8, npa_val, npa_grade, npa_desc)
        write_metric(10, lev_val, lev_grade, lev_desc)

        # Write Overall Status
        c_ov = ws.cell(row=row, column=12)
        c_ov.value = "CLEAR" if overall_g == 'CLEAR' else overall_g
        c_ov.font = data_bold
        c_ov.border = cell_border
        c_ov.alignment = Alignment(horizontal="center", vertical="center")
        
        overall_fill_map = {
            'CLEAR': fill_clear,
            'THRESHOLD_1': fill_watch,
            'THRESHOLD_2': fill_breach,
            'THRESHOLD_3': fill_breach,
            'PENDING': fill_pending
        }
        c_ov.fill = overall_fill_map.get(overall_g, fill_pending)

        # Apply borders to details
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = cell_border

        ws.row_dimensions[row].height = 22

    # Double underline for summary/bottom row
    bottom_row = start_row + len(data['banks'])
    for c in range(1, 13):
        ws.cell(row=bottom_row, column=c).border = summary_border

    # Disclaimer Block
    disclaimer_row = bottom_row + 2
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row+4, end_column=12)
    disclaimer_cell = ws.cell(row=disclaimer_row, column=1)
    disclaimer_cell.value = (
        "DISCLAIMER: This report is an independent regulatory-analytics recreation mapping public capital "
        "and asset-quality disclosures against the RBI's published Prompt Corrective Action (PCA) guidelines "
        "(Circular RBI/2021-22/118). It does not represent or substitute the Reserve Bank of India's official "
        "supervisory assessment, which incorporates non-public inspection reports, qualitative indicators, "
        "and discretionary judgment.\n\n"
        "DATA SOURCE & QUALITY NOTE: All figures are extracted directly from official public investor publications "
        "and results releases of the respective banks for the quarter ended December 31, 2025 (Q3 FY26). "
        "No metrics have been estimated, interpolated, or forecast. Leverage ratio data is marked as 'Data Pending' "
        "since it was not published in the standard Q3 FY26 datasets. Metrics have not been independently audited or verified."
    )
    disclaimer_cell.font = disclaimer_font
    disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid using merged cell width calculation
            if cell.coordinate in [f"A{disclaimer_row}", f"A1", f"A2"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Specific tweaks to column dimensions to avoid truncation
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['K'].width = 18

    # Create Charts
    # Chart 1: CRAR Adequacy Comparison
    chart_crar = BarChart()
    chart_crar.type = "col"
    chart_crar.style = 10
    chart_crar.title = "Capital Adequacy Comparison (CRAR %)"
    chart_crar.y_axis.title = "CRAR Ratio"
    chart_crar.x_axis.title = "Bank Tickers"
    
    data_ref_crar = Reference(ws, min_col=4, min_row=5, max_row=bottom_row-1)
    cats_ref = Reference(ws, min_col=2, min_row=6, max_row=bottom_row-1)
    
    chart_crar.add_data(data_ref_crar, titles_from_data=True)
    chart_crar.set_categories(cats_ref)
    chart_crar.legend = None # No legend needed for single series
    chart_crar.width = 14
    chart_crar.height = 7
    ws.add_chart(chart_crar, "A16")

    # Chart 2: Net NPA Stress Comparison
    chart_npa = BarChart()
    chart_npa.type = "col"
    chart_npa.style = 13
    chart_npa.title = "Asset Quality comparison (Net NPA %)"
    chart_npa.y_axis.title = "Net NPA Ratio"
    chart_npa.x_axis.title = "Bank Tickers"
    
    data_ref_npa = Reference(ws, min_col=8, min_row=5, max_row=bottom_row-1)
    
    chart_npa.add_data(data_ref_npa, titles_from_data=True)
    chart_npa.set_categories(cats_ref)
    chart_npa.legend = None
    chart_npa.width = 14
    chart_npa.height = 7
    ws.add_chart(chart_npa, "G16")

    # Save Excel
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    print(f"Saving compiled spreadsheet to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Report compiled successfully!")

if __name__ == "__main__":
    main()
